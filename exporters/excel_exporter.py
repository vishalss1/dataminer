from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from models.product import Product
from utils.logger import setup_logger

logger = setup_logger(__name__)

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

ROW_ALTERNATE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ROW_NORMAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

CELL_FONT = Font(name="Calibri", size=11)
CELL_ALIGNMENT = Alignment(vertical="center")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CURRENCY_FORMAT = '#,##0.00"€"'  # will be overridden per currency

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")


class ExcelExporter:
    CURRENCY_FORMATS = {
        "USD": '$#,##0.00',
        "EUR": '#,##0.00" €"',
        "GBP": '"£"#,##0.00',
        "INR": '"₹"#,##0.00',
    }

    def __init__(self, currency: str = "USD"):
        self.currency = currency

    def export(self, products: List[Product], filepath: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = [
            "Product Name", "Price", "Rating", "Reviews",
            "Availability", "URL", "Image URL", "Source", "Scraped At",
        ]

        self._write_headers(ws, headers)
        self._write_data(ws, products, headers)
        self._apply_formatting(ws, headers, products)
        self._add_summary_sheet(wb, products)

        wb.save(filepath)
        logger.info(f"Exported {len(products)} products to {filepath}")
        return filepath

    def _write_headers(self, ws, headers: List[str]):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    def _write_data(self, ws, products: List[Product], headers: List[str]):
        for row_idx, product in enumerate(products, 2):
            data = product.to_dict()
            row_fill = ROW_ALTERNATE_FILL if row_idx % 2 == 0 else ROW_NORMAL_FILL

            values = [
                product.name,
                product.price,
                product.rating if product.rating is not None else "",
                product.review_count if product.review_count is not None else "",
                product.availability,
                product.url,
                product.image_url,
                product.source,
                product.scraped_at,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = CELL_FONT
                cell.fill = row_fill
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

                if col_idx == 2:
                    cell.number_format = self.CURRENCY_FORMATS.get(
                        self.currency, '#,##0.00'
                    )
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                if col_idx in (3, 4):
                    cell.alignment = CENTER_ALIGNMENT

                if col_idx == 6 and product.url:
                    cell.font = LINK_FONT
                    cell.hyperlink = product.url

    def _apply_formatting(self, ws, headers: List[str], products: List[Product]):
        ws.freeze_panes = "A2"

        ws.auto_filter.ref = ws.dimensions

        for col_idx, header in enumerate(headers, 1):
            max_length = len(header)
            for row in ws.iter_rows(
                min_col=col_idx, max_col=col_idx,
                min_row=2, max_row=len(products) + 1,
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 3, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        ws.row_dimensions[1].height = 22

    def _add_summary_sheet(self, wb: Workbook, products: List[Product]):
        ws = wb.create_sheet("Summary")
        summary_data = [
            ("Metric", "Value"),
            ("Total Products", len(products)),
            ("Export Date", products[0].scraped_at if products else "N/A"),
            ("Source", products[0].source if products else "N/A"),
            ("", ""),
            ("Price Statistics", ""),
        ]

        prices = [p.price for p in products if p.price]
        if prices:
            summary_data.append(("Average Price", f"${sum(prices) / len(prices):.2f}"))
            summary_data.append(("Min Price", f"${min(prices):.2f}"))
            summary_data.append(("Max Price", f"${max(prices):.2f}"))

        has_ratings = [p.rating for p in products if p.rating is not None]
        if has_ratings:
            summary_data.append(("", ""))
            summary_data.append(("Rating Statistics", ""))
            summary_data.append(("Average Rating", f"{sum(has_ratings) / len(has_ratings):.2f} / 5"))

        for row_idx, (metric, value) in enumerate(summary_data, 1):
            cell_a = ws.cell(row=row_idx, column=1, value=metric)
            cell_b = ws.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                cell_a.font = HEADER_FONT
                cell_a.fill = HEADER_FILL
                cell_b.font = HEADER_FONT
                cell_b.fill = HEADER_FILL
            else:
                cell_a.font = Font(name="Calibri", size=11, bold=True)
                cell_b.font = CELL_FONT
            cell_a.border = THIN_BORDER
            cell_b.border = THIN_BORDER

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.freeze_panes = "A2"
